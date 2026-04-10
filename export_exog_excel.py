"""
Export all exogenous factor datasets used in the FCL forecasting models
into separate, well-labelled Excel files with metadata sheets.
"""
import os, pickle
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

BASE   = '/home/ubuntu/fcl_forecast'
DATA   = os.path.join(BASE, 'data')
OUT    = '/home/ubuntu/fcl_forecast_exog_data'
os.makedirs(OUT, exist_ok=True)

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY   = '1F3864'
BLUE   = '2E75B6'
LBLUE  = 'D6E4F0'
WHITE  = 'FFFFFF'
GREY   = 'F2F2F2'
AMBER  = 'FFC000'
GREEN  = '70AD47'
RED    = 'FF0000'

def style_header_row(ws, row, cols, fill_hex=NAVY, font_hex=WHITE, bold=True):
    fill = PatternFill('solid', fgColor=fill_hex)
    font = Font(color=font_hex, bold=bold, size=11)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        fill = PatternFill('solid', fgColor=GREY if r % 2 == 0 else WHITE)
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 30)

def add_metadata_sheet(wb, info_dict):
    ws = wb.create_sheet('Metadata', 0)
    ws.sheet_view.showGridLines = False
    # Title
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = info_dict.get('title', 'Dataset Metadata')
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    rows = [
        ('Field', 'Value', ''),
        ('Dataset Name', info_dict.get('name', ''), ''),
        ('Source', info_dict.get('source', ''), ''),
        ('Source URL', info_dict.get('url', ''), ''),
        ('Series / Ticker', info_dict.get('series', ''), ''),
        ('Frequency', info_dict.get('frequency', 'Monthly'), ''),
        ('Date Range', info_dict.get('date_range', ''), ''),
        ('Unit', info_dict.get('unit', ''), ''),
        ('Role in Model', info_dict.get('role', ''), ''),
        ('Used in SARIMAX', info_dict.get('sarimax', 'Yes'), ''),
        ('Used in XGBoost', info_dict.get('xgboost', 'Yes (direct + lag-1)'), ''),
        ('Notes', info_dict.get('notes', ''), ''),
    ]
    for i, (k, v, _) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = PatternFill('solid', fgColor=LBLUE)
        ws.cell(row=i, column=2, value=v)
        if i == 3:  # header row
            style_header_row(ws, i, 2, fill_hex=BLUE)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 55

def write_series_sheet(wb, df_col, sheet_name, col_label, fmt='0.000'):
    """Write a single time series column to a sheet with date + value + MoM change."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Header
    headers = ['Date', col_label, 'MoM Change', 'MoM % Change', 'YoY % Change']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    series = df_col.dropna().sort_index()
    mom    = series.diff()
    mom_pct = series.pct_change() * 100
    yoy_pct = series.pct_change(12) * 100

    for r, (idx, val) in enumerate(series.items(), start=2):
        ws.cell(row=r, column=1, value=idx.strftime('%Y-%m-%d'))
        ws.cell(row=r, column=2, value=round(float(val), 6))
        m = mom.get(idx)
        mp = mom_pct.get(idx)
        yp = yoy_pct.get(idx)
        ws.cell(row=r, column=3, value=round(float(m), 6) if pd.notna(m) else None)
        ws.cell(row=r, column=4, value=round(float(mp), 3) if pd.notna(mp) else None)
        ws.cell(row=r, column=5, value=round(float(yp), 3) if pd.notna(yp) else None)

    style_data_rows(ws, 2, r, len(headers))
    auto_col_width(ws)
    return ws

# ══════════════════════════════════════════════════════════════════════════════
# LOAD THE MASTER EXOG DATASET
# ══════════════════════════════════════════════════════════════════════════════
exog = pd.read_parquet(os.path.join(DATA, 'exog_features_antarctica.parquet'))
exog.index = pd.to_datetime(exog.index)
exog = exog.sort_index()

# Also load the original features_monthly for extra columns (mfg, CLI etc.)
feat = pd.read_csv(os.path.join(DATA, 'features_monthly.csv'),
                   index_col=0, parse_dates=True).sort_index()

print(f"Exog shape: {exog.shape}, range: {exog.index[0].date()} – {exog.index[-1].date()}")
print(f"Features_monthly shape: {feat.shape}, range: {feat.index[0].date()} – {feat.index[-1].date()}")
print("Exog columns:", list(exog.columns))
print("Features columns:", list(feat.columns))

# ══════════════════════════════════════════════════════════════════════════════
# DEFINE EACH FACTOR WITH METADATA
# ══════════════════════════════════════════════════════════════════════════════
factors = [
    {
        'filename': '01_Brent_Crude_Oil.xlsx',
        'title': 'Brent Crude Oil Price (USD/bbl)',
        'name': 'Brent Crude Oil',
        'source': 'U.S. Energy Information Administration (EIA) via FRED API',
        'url': 'https://fred.stlouisfed.org/series/DCOILBRENTEU',
        'series': 'FRED: DCOILBRENTEU (monthly average of daily spot prices)',
        'frequency': 'Monthly (averaged from daily)',
        'unit': 'USD per barrel',
        'role': 'Primary cost driver — fuel surcharges and carrier operating costs',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Yes — brent_crude (current) + brent_crude_l1 (1-month lag)',
        'notes': 'Hormuz toggle in dashboard adds +40% to this value as an oil-shock proxy. '
                 'Collected via FRED API (free, no key required for basic access).',
        'col': 'brent_crude',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'Brent Crude (USD/bbl)',
    },
    {
        'filename': '02_USD_CNY_Exchange_Rate.xlsx',
        'title': 'USD / CNY Exchange Rate',
        'name': 'USD/CNY Exchange Rate',
        'source': 'Board of Governors of the Federal Reserve via FRED API',
        'url': 'https://fred.stlouisfed.org/series/DEXCHUS',
        'series': 'FRED: DEXCHUS (monthly average of daily rates)',
        'frequency': 'Monthly (averaged from daily)',
        'unit': 'Chinese Yuan Renminbi per 1 US Dollar',
        'role': 'Currency driver — affects USD-denominated rates on China-origin lanes',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Yes — usdcny (current) + usdcny_l1 (1-month lag)',
        'notes': 'Higher USD/CNY = weaker CNY. Collected via FRED API (free).',
        'col': 'usdcny',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'USD/CNY Rate',
    },
    {
        'filename': '03_US_Industrial_Production.xlsx',
        'title': 'US Industrial Production Index',
        'name': 'US Industrial Production Index',
        'source': 'Federal Reserve Board via FRED API',
        'url': 'https://fred.stlouisfed.org/series/INDPRO',
        'series': 'FRED: INDPRO (seasonally adjusted, 2017=100)',
        'frequency': 'Monthly',
        'unit': 'Index (2017 = 100)',
        'role': 'US demand proxy — higher production = stronger import demand',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Yes — us_indpro (current) + us_indpro_l1 (1-month lag)',
        'notes': 'Seasonally adjusted. Collected via FRED API (free).',
        'col': 'us_indpro',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'US IndPro Index (2017=100)',
    },
    {
        'filename': '04_US_CFNAI.xlsx',
        'title': 'Chicago Fed National Activity Index (CFNAI)',
        'name': 'Chicago Fed National Activity Index (CFNAI)',
        'source': 'Federal Reserve Bank of Chicago via FRED API',
        'url': 'https://fred.stlouisfed.org/series/CFNAI',
        'series': 'FRED: CFNAI',
        'frequency': 'Monthly',
        'unit': 'Index (0 = historical average growth; negative = below-trend)',
        'role': 'Broad US economic activity gauge — leading indicator of trade volumes',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Yes — us_cfnai (current) + us_cfnai_l1 (1-month lag)',
        'notes': 'Designed to have mean 0 and std dev 1. Values below −0.70 '
                 'historically associated with recessions. Collected via FRED API (free).',
        'col': 'us_cfnai',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'CFNAI',
    },
    {
        'filename': '05_China_Exports.xlsx',
        'title': 'China Total Exports (USD)',
        'name': 'China Total Exports',
        'source': 'OECD SDMX REST API (free, no key required)',
        'url': 'https://stats.oecd.org/SDMX-JSON/data/MEI_BOP6/CHN.B6CRSE01.CXCU.M',
        'series': 'OECD MEI BOP6: China Exports of Goods and Services (current USD)',
        'frequency': 'Monthly',
        'unit': 'USD (current prices)',
        'role': 'Supply-side driver — higher exports = more container demand from China',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Not used as separate XGBoost feature (captured by rate lags)',
        'notes': 'Collected via OECD SDMX REST API (completely free, no registration). '
                 'Values in USD current prices.',
        'col': 'china_exports',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'China Exports (USD)',
    },
    {
        'filename': '06_BDRY_ETF_Dry_Bulk_Proxy.xlsx',
        'title': 'BDRY ETF — Dry Bulk Shipping Proxy',
        'name': 'BDRY ETF (Breakwave Dry Bulk Shipping ETF)',
        'source': 'Yahoo Finance via yfinance Python library (free)',
        'url': 'https://finance.yahoo.com/quote/BDRY',
        'series': 'Ticker: BDRY (monthly closing price)',
        'frequency': 'Monthly (last trading day close)',
        'unit': 'USD per ETF unit',
        'role': 'Shipping capacity proxy — tracks Baltic Dry Index (BDI) futures; '
                'rising BDRY = tightening dry bulk capacity = upward pressure on container rates',
        'sarimax': 'Yes — direct coefficient in SARIMAX exog matrix',
        'xgboost': 'Yes — bdry_etf (current) + bdry_etf_l1 (1-month lag)',
        'notes': 'Used as a free, real-time proxy for the Baltic Dry Index (BDI). '
                 'BDRY tracks near-dated BDI futures contracts. Collected via yfinance (free).',
        'col': 'bdry_etf',
        'source_df': exog,
        'sheet': 'Monthly Data',
        'col_label': 'BDRY ETF (USD)',
    },
]

# Dummy / binary event variables
dummy_factors = [
    {
        'col': 'dummy_covid',
        'label': 'COVID Demand Shock',
        'desc': 'Binary dummy = 1 during COVID demand collapse period (Mar 2020 – Jun 2020)',
        'notes': 'Hand-coded binary event dummy. Captures the sharp demand drop at the onset of COVID-19.',
    },
    {
        'col': 'dummy_supply_crunch',
        'label': 'Container Supply Crunch',
        'desc': 'Binary dummy = 1 during the 2021–2022 container supply crunch (Jul 2021 – Dec 2022)',
        'notes': 'Hand-coded binary event dummy. Captures the extraordinary rate spike driven by port congestion, '
                 'equipment shortages, and post-COVID demand surge.',
    },
    {
        'col': 'dummy_ukraine',
        'label': 'Ukraine / Black Sea War',
        'desc': 'Binary dummy = 1 from Feb 2022 onwards (Russia-Ukraine war)',
        'notes': 'Hand-coded binary event dummy. Scaled by lane-specific Ukraine impact matrix '
                 '(Europe 80%, Middle East 40%, East Asia 20%, Americas 10%).',
    },
    {
        'col': 'dummy_red_sea',
        'label': 'Red Sea / Suez Disruption',
        'desc': 'Binary dummy = 1 from Dec 2023 onwards (Houthi attacks on Red Sea shipping)',
        'notes': 'Hand-coded binary event dummy. Scaled by lane-specific Red Sea impact matrix '
                 '(Europe 100%, Middle East 90%, South Asia 70%, SE Asia 50%, East Asia 40%, Americas 15%).',
    },
    {
        'col': 'dummy_hormuz',
        'label': 'Strait of Hormuz Risk',
        'desc': 'Binary dummy = 0 in training data (no full closure in sample period). '
                'In dashboard: activates a +40% Brent crude override (oil shock mechanism).',
        'notes': 'Modelled as an indirect oil-price shock rather than a direct route disruption. '
                 'When toggled ON in the dashboard, Brent crude is multiplied by 1.40 before '
                 'being fed to both SARIMAX and XGBoost.',
    },
]

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 1: Individual factor files (continuous variables)
# ══════════════════════════════════════════════════════════════════════════════
for fac in factors:
    print(f"Writing {fac['filename']} ...")
    wb = __import__('openpyxl').Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    # Metadata sheet
    add_metadata_sheet(wb, fac)

    # Data sheet
    series = fac['source_df'][fac['col']]
    write_series_sheet(wb, series, fac['sheet'], fac['col_label'])

    wb.save(os.path.join(OUT, fac['filename']))

print("Continuous factor files done.")

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 2: Event dummies combined file
# ══════════════════════════════════════════════════════════════════════════════
print("Writing 07_Event_Dummies.xlsx ...")
import openpyxl
wb_d = openpyxl.Workbook()
del wb_d[wb_d.sheetnames[0]]

# Metadata sheet
meta_d = {
    'title': 'Event Dummy Variables',
    'name': 'Geopolitical & Structural Event Dummies',
    'source': 'Hand-coded binary indicators (no external API)',
    'url': 'N/A — researcher-defined based on publicly documented events',
    'series': 'dummy_covid, dummy_supply_crunch, dummy_ukraine, dummy_red_sea, dummy_hormuz, '
              'shock_covid_spike, shock_post_crash',
    'frequency': 'Monthly',
    'unit': 'Binary (0 or 1)',
    'role': 'Capture structural breaks and geopolitical shocks not reflected in continuous variables',
    'sarimax': 'Yes — all dummies included in SARIMAX exog matrix',
    'xgboost': 'Yes — all dummies + lag-1 versions included in XGBoost feature matrix',
    'notes': 'Hormuz dummy is 0 throughout the training period (no full closure occurred). '
             'In the live dashboard it activates a +40% Brent crude override instead of a direct dummy.',
}
add_metadata_sheet(wb_d, meta_d)

# Combined data sheet
ws_d = wb_d.create_sheet('All Dummies')
ws_d.sheet_view.showGridLines = False

dummy_cols = ['dummy_covid', 'dummy_supply_crunch', 'dummy_ukraine',
              'dummy_red_sea', 'dummy_hormuz']
headers = ['Date'] + [
    'COVID Shock\n(Mar–Jun 2020)',
    'Supply Crunch\n(Jul 2021–Dec 2022)',
    'Ukraine War\n(Feb 2022+)',
    'Red Sea\n(Dec 2023+)',
    'Hormuz Risk\n(0 in training)',
]
for c, h in enumerate(headers, 1):
    ws_d.cell(row=1, column=c, value=h)
style_header_row(ws_d, 1, len(headers))
ws_d.row_dimensions[1].height = 35

dummy_df = exog[dummy_cols].sort_index()
for r, (idx, row) in enumerate(dummy_df.iterrows(), start=2):
    ws_d.cell(row=r, column=1, value=idx.strftime('%Y-%m-%d'))
    for c, col in enumerate(dummy_cols, start=2):
        cell = ws_d.cell(row=r, column=c, value=int(row[col]))
        if int(row[col]) == 1:
            cell.fill = PatternFill('solid', fgColor='FFD966')  # amber for active
            cell.font = Font(bold=True)

style_data_rows(ws_d, 2, r, len(headers))
auto_col_width(ws_d)

# Individual dummy description sheets
for df_info in dummy_factors:
    safe_title = df_info['label'].replace('/', '-').replace('\\', '-')[:31]
    ws_i = wb_d.create_sheet(safe_title)
    ws_i.sheet_view.showGridLines = False
    # Description box
    ws_i.merge_cells('A1:B1')
    ws_i['A1'].value = df_info['label']
    ws_i['A1'].font = Font(bold=True, size=12, color=WHITE)
    ws_i['A1'].fill = PatternFill('solid', fgColor=BLUE)
    ws_i['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_i.row_dimensions[1].height = 25

    ws_i.cell(row=2, column=1, value='Description').font = Font(bold=True)
    ws_i.cell(row=2, column=2, value=df_info['desc'])
    ws_i.cell(row=3, column=1, value='Notes').font = Font(bold=True)
    ws_i.cell(row=3, column=2, value=df_info['notes'])
    ws_i.column_dimensions['A'].width = 18
    ws_i.column_dimensions['B'].width = 70

    ws_i.cell(row=5, column=1, value='Date').font = Font(bold=True)
    ws_i.cell(row=5, column=2, value='Value (0/1)').font = Font(bold=True)
    style_header_row(ws_i, 5, 2)

    col_data = exog[df_info['col']].sort_index()
    for rr, (idx, val) in enumerate(col_data.items(), start=6):
        ws_i.cell(row=rr, column=1, value=idx.strftime('%Y-%m-%d'))
        cell = ws_i.cell(row=rr, column=2, value=int(val))
        if int(val) == 1:
            cell.fill = PatternFill('solid', fgColor='FFD966')
            cell.font = Font(bold=True)

wb_d.save(os.path.join(OUT, '07_Event_Dummies.xlsx'))
print("Event dummies file done.")

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 3: Master combined file — all exog in one workbook
# ══════════════════════════════════════════════════════════════════════════════
print("Writing 00_Master_All_Exog_Factors.xlsx ...")
wb_m = openpyxl.Workbook()
del wb_m[wb_m.sheetnames[0]]

# Metadata
meta_m = {
    'title': 'FCL Forecast — All Exogenous Factors (Master)',
    'name': 'Master Exogenous Factor Dataset',
    'source': 'FRED API (Brent, USD/CNY, IndPro, CFNAI) · OECD SDMX REST (China Exports) · '
              'Yahoo Finance/yfinance (BDRY ETF) · Hand-coded (Event Dummies)',
    'url': 'See individual factor sheets for source URLs',
    'series': 'brent_crude, usdcny, us_indpro, us_cfnai, china_exports, bdry_etf, '
              'dummy_covid, dummy_supply_crunch, dummy_ukraine, dummy_red_sea, dummy_hormuz',
    'frequency': 'Monthly',
    'unit': 'Mixed — see individual sheets',
    'role': 'Full exogenous feature matrix used in SARIMAX training (Path 1: 12 cols, Path 2: 10 cols)',
    'sarimax': 'Yes — all 12 columns (Path 1) / 10 columns (Path 2)',
    'xgboost': 'Yes — all continuous vars + lag-1 versions + rate lags + rolling means + month dummies + sar_pred',
    'notes': 'Date range: Jul 2019 – Apr 2026 (88 monthly observations). '
             'Training window: Jul 2019 – Dec 2024. Jan 2025–Apr 2026 used for validation.',
}
add_metadata_sheet(wb_m, meta_m)

# Combined data sheet — all columns
ws_all = wb_m.create_sheet('All Factors (Wide)')
ws_all.sheet_view.showGridLines = False

all_cols = list(exog.columns)
col_labels = {
    'brent_crude':         'Brent Crude\n(USD/bbl)',
    'usdcny':              'USD/CNY\nRate',
    'us_indpro':           'US IndPro\nIndex (2017=100)',
    'us_cfnai':            'CFNAI',
    'china_exports':       'China Exports\n(USD)',
    'bdry_etf':            'BDRY ETF\n(USD)',
    'dummy_covid':         'COVID\nShock',
    'dummy_supply_crunch': 'Supply\nCrunch',
    'dummy_ukraine':       'Ukraine\nWar',
    'dummy_red_sea':       'Red Sea\nDisruption',
    'dummy_hormuz':        'Hormuz\nRisk',
}
headers_all = ['Date'] + [col_labels.get(c, c) for c in all_cols]
for col_i, h in enumerate(headers_all, 1):
    ws_all.cell(row=1, column=col_i, value=h)
style_header_row(ws_all, 1, len(headers_all))
ws_all.row_dimensions[1].height = 35

for r, (idx, row) in enumerate(exog.iterrows(), start=2):
    ws_all.cell(row=r, column=1, value=idx.strftime('%Y-%m-%d'))
    for c_i, col in enumerate(all_cols, start=2):
        val = row[col]
        cell = ws_all.cell(row=r, column=c_i, value=round(float(val), 6) if pd.notna(val) else None)
        # Highlight active dummies
        if col.startswith('dummy_') and int(val) == 1:
            cell.fill = PatternFill('solid', fgColor='FFD966')
            cell.font = Font(bold=True)

style_data_rows(ws_all, 2, r, len(headers_all))
auto_col_width(ws_all)

# Individual sheets for each continuous factor
for fac in factors:
    write_series_sheet(wb_m, exog[fac['col']], fac['col'][:31], fac['col_label'])

wb_m.save(os.path.join(OUT, '00_Master_All_Exog_Factors.xlsx'))
print("Master file done.")

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 4: Extended features (from features_monthly.csv — extra indicators)
# ══════════════════════════════════════════════════════════════════════════════
print("Writing 08_Extended_Features_Pipeline.xlsx ...")
wb_e = openpyxl.Workbook()
del wb_e[wb_e.sheetnames[0]]

meta_e = {
    'title': 'Extended Feature Pipeline (features_monthly.csv)',
    'name': 'Extended Feature Set — Data Pipeline Output',
    'source': 'FRED API · OECD SDMX REST · Yahoo Finance · Drewry WCI (compiled)',
    'url': 'See individual factor sheets',
    'series': 'brent_crude_usd, usdcny, us_indpro, us_mfg_emp, us_cfnai, us_mfg_orders, '
              'china_exports_usd, us_imports_china, china_cli, us_cli, bdry_etf, wci_composite + lags/MAs',
    'frequency': 'Monthly',
    'unit': 'Mixed',
    'role': 'Full feature set generated by data_pipeline.py — includes lags, moving averages, '
            'and log-returns. Subset used for final model training.',
    'sarimax': 'Subset used (see exog_features_antarctica.parquet)',
    'xgboost': 'Subset used (see model bundle xgb_feat_cols)',
    'notes': 'Date range: Jan 2021 – Dec 2024 (48 months). This is the raw pipeline output '
             'before the Antarctica-specific feature engineering step.',
}
add_metadata_sheet(wb_e, meta_e)

# Raw data sheet
ws_e = wb_e.create_sheet('All Features (Wide)')
ws_e.sheet_view.showGridLines = False

feat_cols = list(feat.columns)
for c_i, h in enumerate(['Date'] + feat_cols, 1):
    ws_e.cell(row=1, column=c_i, value=h)
style_header_row(ws_e, 1, len(feat_cols) + 1)

for r, (idx, row) in enumerate(feat.iterrows(), start=2):
    ws_e.cell(row=r, column=1, value=idx.strftime('%Y-%m-%d'))
    for c_i, col in enumerate(feat_cols, start=2):
        val = row[col]
        ws_e.cell(row=r, column=c_i,
                  value=round(float(val), 6) if pd.notna(val) else None)

style_data_rows(ws_e, 2, r, len(feat_cols) + 1)
auto_col_width(ws_e)

# Key individual series sheets from features_monthly
extra_series = {
    'wci_composite':      ('Drewry WCI Composite\n(USD/40ft)', 'Drewry WCI Composite Rate'),
    'us_mfg_emp':         ('US Mfg Employment\n(thousands)', 'US Manufacturing Employment'),
    'us_mfg_orders':      ('US Mfg Orders\n(USD millions)', 'US Manufacturing New Orders'),
    'us_imports_china':   ('US Imports from China\n(USD)', 'US Imports from China'),
    'china_cli':          ('China CLI', 'China Composite Leading Indicator (OECD)'),
    'us_cli':             ('US CLI', 'US Composite Leading Indicator (OECD)'),
}
for col, (label, sheet_title) in extra_series.items():
    if col in feat.columns:
        write_series_sheet(wb_e, feat[col], sheet_title[:31], label)

wb_e.save(os.path.join(OUT, '08_Extended_Features_Pipeline.xlsx'))
print("Extended features file done.")

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
files = sorted(os.listdir(OUT))
print(f"\n{'='*60}")
print(f"Exported {len(files)} files to {OUT}:")
for f in files:
    size = os.path.getsize(os.path.join(OUT, f))
    print(f"  {f}  ({size/1024:.1f} KB)")
